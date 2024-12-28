import os
import re
import pandas as pd
from openpyxl.reader.excel import load_workbook
from pandas import Series

data = pd.read_csv("data.csv")

question_frequencies = {}
# "", "", "", "", ""
# 0: single-selection, 1: multi-selection

question_answers_i = {
    f"question_1": {"question":"Yaşınız nedir?", "type":0, "answers":["18-24", "25-34", "35-44", "45-54", "55 ve üzeri"]},
    f"question_2": {"question":"Cinsiyetiniz nedir?", "type":0, "answers":["Kadın", "Erkek"]},
    f"question_3": {"question":"Eğitim durumunuz nedir?", "type":0, "answers":["Ortaokul", "Lise", "Üniversite", "Yüksek Lisans", "Doktora"]},
    f"question_4": {"question":"Bulunduğunuz bölge hangisidir?", "type":0, "answers":["Ege", "Marmara", "İç Anadolu", "Akdeniz", "Karadeniz", "Güneydoğu Anadolu"]},
    f"question_5": {"question":"Çalışma durumunuz nedir?", "type":0, "answers":["Çalışıyorum", "Çalışmıyorum", "Öğrenci", "Emekli"]},
    f"question_6": {"question":"Aylık gelir düzeyiniz nedir?", "type":0, "answers":["0-5.000 TL", "5.000-10.000 TL", "10.000-20.000 TL", "20.000-30.000 TL", "30.000 TL ve üzeri"]},
    f"question_7": {"question":"STK’ların toplum üzerindeki etkisini nasıl değerlendiriyorsunuz?", "type":0, "answers":["Hiç etkili değil", "Az etkili", "Orta düzeyde", "Etkili", "Çok etkili"]},
    f"question_8": {"question":"STK’ların finansman sağlama konusunda en büyük zorluğu nedir?", "type":0, "answers":["Kaynak yetersizliği", "Bağış eksikliği", "Devlet desteği azlığı", "Yönetim sorunları"]},
    f"question_9": {"question":"STK’ların gücünü arttırmak için hangi stratejiler izlenmeli?", "type":1, "answers":["Daha fazla reklam ve tanıtım yapmalı", "İş birliği artırma", "Dijital katılımı iyileştirilmeli", "Öğrenci liderliğindeki girişimleri kolaylaştırmalı", "Eğitim ortaklıkları geliştirilmeli", "Kampüs varlığı güçlendirilmeli", "Liderliği ve yeniliği teşvik etmeli", "Gönüllüleri tanıtmalı ve motive etmeli", "Esnek gönüllü seçenekleri sunmalı"]},
    f"question_10": {"question":"Topluma fayda sağlamak adına STK'ların etkisini nasıl görüyorsunuz?", "type":0, "answers":["Hiç etkili değil", "Az etkili", "Orta düzeyde", "Etkili", "Çok etkili"]},
    f"question_11": {"question":"STK’lar gençlerin ilgisini çekmek için ne yapmalı?", "type":1, "answers":["Dijital katılım iyileştirilmeli", "Öğrenci liderliğindeki girişimler kolaylaştırılmalı", "Eğitim ortaklıkları geliştirilmeli", "Kampüs varlığı güçlendirilmeli", "Liderliği ve yeniliği teşvik etmeli", "Esnek gönüllü seçenekleri sunulmalı", "Sosyal medyada daha aktif olunmalı", "Etkinlik düzenlenmeli", "Gençlik programları oluşturulmalı"]},
    f"question_12": {"question":"Toplumun STK’lara olan güvenini nasıl değerlendiriyorsunuz?", "type":0, "answers":["Hiç güvenmiyor", "Az güveniyor", "Orta seviyede", "Güveniyor", "Çok güveniyor"]},
    f"question_13": {"question":"STK'ların kurumsal firmalarla ortak projeler yürütmesini destekliyor musunuz?", "type":0, "answers":["Kararsızım", "Hayır", "Evet"]},
    f"question_14": {"question":"STK'ların uluslararası iş birlikleri hakkında ne düşünüyorsunuz?", "type":0, "answers":["Hiç önemli değil", "Az önemli", "Orta düzeyde", "Önemli", "Çok önemli"]},
    f"question_15": {"question":"STK'ların toplumsal bilinç oluşturma konusunda başarısı nedir?", "type":0, "answers":["Çok başarısız", "Başarısız", "Orta düzeyde", "Başarılı", "Çok başarılı"]},
    f"question_16": {"question":"Sizce STK'lar hangi alanlarda daha fazla faaliyet göstermeli?", "type":1, "answers":["Çevre", "Eğitim", "Sağlık", "İnsan hakları", "Hayvan hakları"]},
    f"question_17": {"question":"STK'ların bağımsız ve tarafsız olması gerektiğine inanıyor musunuz?", "type":0, "answers":["Kararsızım", "Hayır", "Evet"]},
    f"question_18": {"question":"STK'ların devlet desteklerinden faydalanması hakkında düşünceniz nedir?", "type":0, "answers":["Kararsızım", "Desteklemiyorum", "Kısmen destekliyorum", "Tamamen destekliyorum"]},
    f"question_19": {"question":"STK’ların dijital dünyada daha fazla yer alması gerektiğini düşünüyor musunuz?", "type":0, "answers":["Kararsızım", "Hayır", "Evet", "Kesinlikle evet"]},
    f"question_20": {"question":"Bir STK’ya destek vermeyi tercih ettiğiniz yöntemler hangileridir?", "type":1, "answers":["Maddi bağış", "Gönüllü olarak çalışmak", "Etkinliklere katılmak", "Sosyal medyada paylaşım yapmak", "Bilgi paylaşımı veya önerilerde bulunmak"]},
    f"question_21": {"question":"Gönüllü olarak çalışmayı tercih etmenizde aşağıdakilerden hangisi en etkili faktör olur?", "type":0, "answers":["Yardım etme isteği", "Topluma katkıda bulunma", "Kişisel gelişim", "Sosyal çevre oluşturma", "Mesleki deneyim kazanma"]},
    f"question_22": {"question":"Kanser ile mücadele eden bu kuruluşlardan hangilerini tanıyorsunuz?", "type":1, "answers":["LÖSEV (Lösemili Çocuklar Vakfı)", "KAÇUV (Kanserli Çocuklara Umut Vakfı)", "Türk Kanser Derneği", "KİTVAK (Kemik İliği Transplantasyon ve Onkoloji Merkezi Kurma ve Geliştirme Vakfı)", "Kanser Savaşçıları Derneği", "TÜRKÖK (Türkiye Kök Hücre Koordinasyon Merkezi)"]},
    f"question_23": {"question":"KİTVAK'ı daha önce duymuş muydunuz?", "type":0, "answers":["Hayır", "Evet"]},
    f"question_24": {"question":"Bir önceki soruya ""Evet"" cevabını verdiyseniz KİTVAK'ı nereden duydunuz?", "type":0, "answers":["Sosyal medya (Instagram, Facebook, Twitter, LinkedIn vb.)", "Televizyon veya radyo programı", "Haberler veya basın yayın organları (gazete, dergi, internet haber siteleri)", "KİTVAK'ın düzenlediği etkinlikler (bağış kampanyaları, farkındalık günleri, sosyal sorumluluk projeleri)", "Hastaneler veya sağlık kuruluşları", "Aileden, arkadaşlardan veya tanıdıklardan", "E-posta veya SMS bilgilendirmesi", "KİTVAK’ın kendi internet sitesi", "Diğer sivil toplum kuruluşları veya ortak projeler", "Üniversiteler veya eğitim kurumları", "Hiç duymadım"]},
    f"question_25": {"question":"KİTVAK’ın logosunu daha önce gördünüz mü?", "type":0, "answers":["Evet, hatırlıyorum ve tanıyorum.", "Evet, ama nerede gördüğümü hatırlamıyorum.", "Hayır, hiç görmedim."]},
    f"question_26": {"question":"KİTVAK’ın logosunu gördüğünüzde size ne ifade ediyor?", "type":0, "answers":["Sağlık ve şifa", "Dayanışma ve yardımlaşma", "Güven ve profesyonellik", "Eğitim ve farkındalık", "Hiçbir şey çağrıştırmıyor"]},
    f"question_27": {"question":"Sizce bu logo KİTVAK'ın amacını doğrudan aktarıyor mu?", "type":0, "answers":["Hayır, logo KİTVAK'ın amacını tam olarak yansıtmıyor.", "Kısmen, bazı unsurlar açıklayıcı ama daha fazla netlik eklenebilir.", "Evet, logo KİTVAK'ın amacını açıkça yansıtıyor."]},
    f"question_28": {"question":"Sizce KİTVAK’ın logosu yenilenmeli mi?", "type":0, "answers":["Kararsızım", "Hayır", "Evet", "Kesinlikle evet"]},
    f"question_29": {"question":"Siz olsaydınız KİTVAK Derneğinin logosunda hangi sembollerin olmasını daha doğru bulurdunuz?", "type":0, "answers":["Elleri içeren bir sembol (hastalara ve ailelerine verilen desteği temsil eder).", "DNA sarmalı (KİTVAK’ın bilimsel ve tıbbi kimliğini vurgular).", "Bir ev veya çatı figürü (barınma, destek vurgusu ve KİTVAK’ın konukevi projelerine atıfta bulunur).", "Kalp simgesi (sevgi, şefkat ve yaşamı simgelemek için).", "Açan Çiçek (KİTVAK'ın hastalara sunduğu umut ve iyileşme sürecini vurgular)."]},
}

column_index = 1
for column in data.columns[1:]:
    data.rename(columns={column:f"question_{column_index}"}, inplace=True)
    column_index += 1

table: dict = {}

for column in data.columns[1:]:
    table[column] = {}
    question_frequencies = data[column].value_counts()

    for answer in question_answers_i[column]["answers"]:
        if answer not in question_frequencies.index:
            question_frequencies[answer] = 0

    question_frequencies.name = question_answers_i[column]["question"]

    table[column]["frequency"] = question_frequencies
    table[column]["frequency"]["Total"] = table[column]["frequency"].sum()


for key in table.keys():
    percentages = {}
    valid_percentages = {}
    cumulative_percentages = {}
    cumulative_percentage = 0
    total_percent = 0
    for answer in table[key]["frequency"].keys():
        if answer == "Total":
            continue
        percent = round((table[key]["frequency"][answer] * 100) / table[key]["frequency"]["Total"], 2)
        percentages[answer] = percent
        total_percent += percent
    table[key]["percentages"] = Series(percentages)
    table[key]["percentages"]["Total"] = round(total_percent, 2)

    percentage_remaining = 100 - total_percent
    valid_percentages = percentages
    first_key = next(iter(percentages))
    valid_percentages[first_key] += percentage_remaining
    valid_percentages[first_key] = round(valid_percentages[first_key], 2)
    table[key]["valid_percentages"] = Series(valid_percentages)
    table[key]["valid_percentages"]["Total"] = 100


    for answer in table[key]["valid_percentages"].keys():
        if answer == "Total":
            continue
        cumulative_percentage += valid_percentages[answer]
        cumulative_percentages[answer] = round(cumulative_percentage, 2)

    table[key]["cumulative_percentages"] = cumulative_percentages

def sanitize_sheet_name(name):
    return re.sub(r'[\/:*?"<>|]', '', name)[:31]

def with_template(template_path, output_path):
    template_file = template_path

    for key in table.keys():
        freq_df = table[key]["frequency"]
        perc_df = table[key]["percentages"]
        val_perc_df = table[key]["valid_percentages"]
        cum_perc_df = table[key]["cumulative_percentages"]

        combined_df = pd.DataFrame({
            "Frequency": freq_df,
            "Percentage": perc_df,
            "Valid Percentage": val_perc_df,
            "Cumulative Percentages": cum_perc_df
        })

        total_row = combined_df.loc["Total"]
        sorted_df = combined_df.drop(index="Total").sort_values(by="Frequency", ascending=False)

        combined_df = pd.concat([sorted_df, total_row.to_frame().T])

        combined_df.index.name = question_answers_i[key]['question']

        sheet_name = sanitize_sheet_name(question_answers_i[key]['question'])

        wb = load_workbook(template_file)

        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        with pd.ExcelWriter(template_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            combined_df.to_excel(writer, sheet_name=sheet_name, startrow=1, index=True)

        wb.save(output_path)


def without_template(output_directory_path):
    for key in table.keys():
        freq_df = table[key]["frequency"]
        perc_df = table[key]["percentages"]
        val_perc_df = table[key]["valid_percentages"]
        cum_perc_df = table[key]["cumulative_percentages"]

        combined_df = pd.DataFrame({
            "Frequency": freq_df,
            "Percentage": perc_df,
            "Valid Percentage": val_perc_df,
            "Cumulative Percentages": cum_perc_df
        })

        total_row = combined_df.loc["Total"]
        sorted_df = combined_df.drop(index="Total").sort_values(by="Frequency", ascending=False)
        combined_df = pd.concat([sorted_df, total_row.to_frame().T])

        combined_df.index.name = question_answers_i[key]['question']
        sheet_name = sanitize_sheet_name(question_answers_i[key]['question'])

        os.makedirs(output_directory_path, exist_ok=True)
        combined_df.to_excel(f"{output_directory_path}/{key}.xlsx", index=True, sheet_name=sheet_name)

with_template("template.xlsx", "output_with_data.xlsx")

without_template(f"{os.getcwd()}/output")




